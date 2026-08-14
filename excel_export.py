from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def export_to_excel(students, student_file):
    """
    Export the current student list to an Excel file.
    """

    if len(students) == 0:
        print("❌ No students available.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = ["Name", "Roll Number", "Subject", "Marks", "Grade"]
    ws.append(headers)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="2C3E50"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
        size=12,
        name="Calibri"
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for student in students:
        ws.append([
            student["name"],
            student["rollno"],
            student["subject"],
            student["marks"],
            student["grade"]
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 4

    ws.freeze_panes = "A2"

    excel_file = student_file.replace(".csv", ".xlsx")

    wb.save(excel_file)

    print(f"✅ Excel file saved as '{excel_file}'")
    print(f"✅ Excel file created successfully! to {excel_file}")